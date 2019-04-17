import openpyxl
import openpyxl.styles

wb = openpyxl.load_workbook(r'E:\lianxi\PY\五金生产效率考核数据记录表.xlsx')
#wb1需读取数据，路径不可改
wb1 = openpyxl.load_workbook(r'E:\机械五金文件\日报考核\五金生产效率考核数据记录表.xlsx')

data7=['=IF(D4>0,D4+E4,F4)', '=IF(D5>0,D5+E5,F5)', '=IF(D6>0,D6+E6,F6)', '=IF(D7>0,D7+E7,F7)', '=IF(D8>0,D8+E8,F8)', '=IF(D9>0,D9+E9,F9)', '=IF(D10>0,D10+E10,F10)', '=IF(D11>0,D11+E11,F11)', '=IF(D12>0,D12+E12,F12)', '=IF(D13>0,D13+E13,F13)', '=IF(D14>0,D14+E14,F14)', '=IF(D15>0,D15+E15,F15)', '=IF(D16>0,D16+E16,F16)', '=IF(D17>0,D17+E17,F17)', '=IF(D18>0,D18+E18,F18)', '=IF(D19>0,D19+E19,F19)', '=IF(D20>0,D20+E20,F20)', '=IF(D21>0,D21+E21,F21)', '=IF(D22>0,D22+E22,F22)', '=IF(D23>0,D23+E23,F23)']
data8=['=IFERROR(AVERAGE(N4,R4,V4,Z4,AD4,AH4,AL4,AP4,AT4,AX4),"")', '=IFERROR(AVERAGE(N5,R5,V5,Z5,AD5,AH5,AL5,AP5,AT5,AX5),"")', '=IFERROR(AVERAGE(N6,R6,V6,Z6,AD6,AH6,AL6,AP6,AT6,AX6),"")', '=IFERROR(AVERAGE(N7,R7,V7,Z7,AD7,AH7,AL7,AP7,AT7,AX7),"")', '=IFERROR(AVERAGE(N8,R8,V8,Z8,AD8,AH8,AL8,AP8,AT8,AX8),"")', '=IFERROR(AVERAGE(N9,R9,V9,Z9,AD9,AH9,AL9,AP9,AT9,AX9),"")', '=IFERROR(AVERAGE(N10,R10,V10,Z10,AD10,AH10,AL10,AP10,AT10,AX10),"")', '=IFERROR(AVERAGE(N11,R11,V11,Z11,AD11,AH11,AL11,AP11,AT11,AX11),"")', '=IFERROR(AVERAGE(N12,R12,V12,Z12,AD12,AH12,AL12,AP12,AT12,AX12),"")', '=IFERROR(AVERAGE(N13,R13,V13,Z13,AD13,AH13,AL13,AP13,AT13,AX13),"")', '=IFERROR(AVERAGE(N14,R14,V14,Z14,AD14,AH14,AL14,AP14,AT14,AX14),"")', '=IFERROR(AVERAGE(N15,R15,V15,Z15,AD15,AH15,AL15,AP15,AT15,AX15),"")', '=IFERROR(AVERAGE(N16,R16,V16,Z16,AD16,AH16,AL16,AP16,AT16,AX16),"")', '=IFERROR(AVERAGE(N17,R17,V17,Z17,AD17,AH17,AL17,AP17,AT17,AX17),"")', '=IFERROR(AVERAGE(N18,R18,V18,Z18,AD18,AH18,AL18,AP18,AT18,AX18),"")', '=IFERROR(AVERAGE(N19,R19,V19,Z19,AD19,AH19,AL19,AP19,AT19,AX19),"")', '=IFERROR(AVERAGE(N20,R20,V20,Z20,AD20,AH20,AL20,AP20,AT20,AX20),"")', '=IFERROR(AVERAGE(N21,R21,V21,Z21,AD21,AH21,AL21,AP21,AT21,AX21),"")', '=IFERROR(AVERAGE(N22,R22,V22,Z22,AD22,AH22,AL22,AP22,AT22,AX22),"")', '=IFERROR(AVERAGE(N23,R23,V23,Z23,AD23,AH23,AL23,AP23,AT23,AX23),"")']
data9=['=SUM(N4,R4,V4,Z4,AD4,AH4,AL4,AP4,AT4,AX4)', '=SUM(N5,R5,V5,Z5,AD5,AH5,AL5,AP5,AT5,AX5)', '=SUM(N6,R6,V6,Z6,AD6,AH6,AL6,AP6,AT6,AX6)', '=SUM(N7,R7,V7,Z7,AD7,AH7,AL7,AP7,AT7,AX7)', '=SUM(N8,R8,V8,Z8,AD8,AH8,AL8,AP8,AT8,AX8)', '=SUM(N9,R9,V9,Z9,AD9,AH9,AL9,AP9,AT9,AX9)', '=SUM(N10,R10,V10,Z10,AD10,AH10,AL10,AP10,AT10,AX10)', '=SUM(N11,R11,V11,Z11,AD11,AH11,AL11,AP11,AT11,AX11)', '=SUM(N12,R12,V12,Z12,AD12,AH12,AL12,AP12,AT12,AX12)', '=SUM(N13,R13,V13,Z13,AD13,AH13,AL13,AP13,AT13,AX13)', '=SUM(N14,R14,V14,Z14,AD14,AH14,AL14,AP14,AT14,AX14)', '=SUM(N15,R15,V15,Z15,AD15,AH15,AL15,AP15,AT15,AX15)', '=SUM(N16,R16,V16,Z16,AD16,AH16,AL16,AP16,AT16,AX16)', '=SUM(N17,R17,V17,Z17,AD17,AH17,AL17,AP17,AT17,AX17)', '=SUM(N18,R18,V18,Z18,AD18,AH18,AL18,AP18,AT18,AX18)', '=SUM(N19,R19,V19,Z19,AD19,AH19,AL19,AP19,AT19,AX19)', '=SUM(N20,R20,V20,Z20,AD20,AH20,AL20,AP20,AT20,AX20)', '=SUM(N21,R21,V21,Z21,AD21,AH21,AL21,AP21,AT21,AX21)', '=SUM(N22,R22,V22,Z22,AD22,AH22,AL22,AP22,AT22,AX22)', '=SUM(N23,R23,V23,Z23,AD23,AH23,AL23,AP23,AT23,AX23)']
data10=['=SUM(M4,Q4,U4,Y4,AC4,AG4,AK4,AO4,AS4,AW4)', '=SUM(M5,Q5,U5,Y5,AC5,AG5,AK5,AO5,AS5,AW5)', '=SUM(M6,Q6,U6,Y6,AC6,AG6,AK6,AO6,AS6,AW6)', '=SUM(M7,Q7,U7,Y7,AC7,AG7,AK7,AO7,AS7,AW7)', '=SUM(M8,Q8,U8,Y8,AC8,AG8,AK8,AO8,AS8,AW8)', '=SUM(M9,Q9,U9,Y9,AC9,AG9,AK9,AO9,AS9,AW9)', '=SUM(M10,Q10,U10,Y10,AC10,AG10,AK10,AO10,AS10,AW10)', '=SUM(M11,Q11,U11,Y11,AC11,AG11,AK11,AO11,AS11,AW11)', '=SUM(M12,Q12,U12,Y12,AC12,AG12,AK12,AO12,AS12,AW12)', '=SUM(M13,Q13,U13,Y13,AC13,AG13,AK13,AO13,AS13,AW13)', '=SUM(M14,Q14,U14,Y14,AC14,AG14,AK14,AO14,AS14,AW14)', '=SUM(M15,Q15,U15,Y15,AC15,AG15,AK15,AO15,AS15,AW15)', '=SUM(M16,Q16,U16,Y16,AC16,AG16,AK16,AO16,AS16,AW16)', '=SUM(M17,Q17,U17,Y17,AC17,AG17,AK17,AO17,AS17,AW17)', '=SUM(M18,Q18,U18,Y18,AC18,AG18,AK18,AO18,AS18,AW18)', '=SUM(M19,Q19,U19,Y19,AC19,AG19,AK19,AO19,AS19,AW19)', '=SUM(M20,Q20,U20,Y20,AC20,AG20,AK20,AO20,AS20,AW20)', '=SUM(M21,Q21,U21,Y21,AC21,AG21,AK21,AO21,AS21,AW21)', '=SUM(M22,Q22,U22,Y22,AC22,AG22,AK22,AO22,AS22,AW22)', '=SUM(M23,Q23,U23,Y23,AC23,AG23,AK23,AO23,AS23,AW23)']
data14=['=IF(L4>0,M4/L4,"")', '=IF(L5>0,M5/L5,"")', '=IF(L6>0,M6/L6,"")', '=IF(L7>0,M7/L7,"")', '=IF(L8>0,M8/L8,"")', '=IF(L9>0,M9/L9,"")', '=IF(L10>0,M10/L10,"")', '=IF(L11>0,M11/L11,"")', '=IF(L12>0,M12/L12,"")', '=IF(L13>0,M13/L13,"")', '=IF(L14>0,M14/L14,"")', '=IF(L15>0,M15/L15,"")', '=IF(L16>0,M16/L16,"")', '=IF(L17>0,M17/L17,"")', '=IF(L18>0,M18/L18,"")', '=IF(L19>0,M19/L19,"")', '=IF(L20>0,M20/L20,"")', '=IF(L21>0,M21/L21,"")', '=IF(L22>0,M22/L22,"")', '=IF(L23>0,M23/L23,"")']
data18=['=IF(P4>0,Q4/P4,"")', '=IF(P5>0,Q5/P5,"")', '=IF(P6>0,Q6/P6,"")', '=IF(P7>0,Q7/P7,"")', '=IF(P8>0,Q8/P8,"")', '=IF(P9>0,Q9/P9,"")', '=IF(P10>0,Q10/P10,"")', '=IF(P11>0,Q11/P11,"")', '=IF(P12>0,Q12/P12,"")', '=IF(P13>0,Q13/P13,"")', '=IF(P14>0,Q14/P14,"")', '=IF(P15>0,Q15/P15,"")', '=IF(P16>0,Q16/P16,"")', '=IF(P17>0,Q17/P17,"")', '=IF(P18>0,Q18/P18,"")', '=IF(P19>0,Q19/P19,"")', '=IF(P20>0,Q20/P20,"")', '=IF(P21>0,Q21/P21,"")', '=IF(P22>0,Q22/P22,"")', '=IF(P23>0,Q23/P23,"")']
data22=['=IF(T4>0,U4/T4,"")', '=IF(T5>0,U5/T5,"")', '=IF(T6>0,U6/T6,"")', '=IF(T7>0,U7/T7,"")', '=IF(T8>0,U8/T8,"")', '=IF(T9>0,U9/T9,"")', '=IF(T10>0,U10/T10,"")', '=IF(T11>0,U11/T11,"")', '=IF(T12>0,U12/T12,"")', '=IF(T13>0,U13/T13,"")', '=IF(T14>0,U14/T14,"")', '=IF(T15>0,U15/T15,"")', '=IF(T16>0,U16/T16,"")', '=IF(T17>0,U17/T17,"")', '=IF(T18>0,U18/T18,"")', '=IF(T19>0,U19/T19,"")', '=IF(T20>0,U20/T20,"")', '=IF(T21>0,U21/T21,"")', '=IF(T22>0,U22/T22,"")', '=IF(T23>0,U23/T23,"")']
data26=['=IF(X4>0,Y4/X4,"")', '=IF(X5>0,Y5/X5,"")', '=IF(X6>0,Y6/X6,"")', '=IF(X7>0,Y7/X7,"")', '=IF(X8>0,Y8/X8,"")', '=IF(X9>0,Y9/X9,"")', '=IF(X10>0,Y10/X10,"")', '=IF(X11>0,Y11/X11,"")', '=IF(X12>0,Y12/X12,"")', '=IF(X13>0,Y13/X13,"")', '=IF(X14>0,Y14/X14,"")', '=IF(X15>0,Y15/X15,"")', '=IF(X16>0,Y16/X16,"")', '=IF(X17>0,Y17/X17,"")', '=IF(X18>0,Y18/X18,"")', '=IF(X19>0,Y19/X19,"")', '=IF(X20>0,Y20/X20,"")', '=IF(X21>0,Y21/X21,"")', '=IF(X22>0,Y22/X22,"")', '=IF(X23>0,Y23/X23,"")']
data30=['=IF(AB4>0,AC4/AB4,"")', '=IF(AB5>0,AC5/AB5,"")', '=IF(AB6>0,AC6/AB6,"")', '=IF(AB7>0,AC7/AB7,"")', '=IF(AB8>0,AC8/AB8,"")', '=IF(AB9>0,AC9/AB9,"")', '=IF(AB10>0,AC10/AB10,"")', '=IF(AB11>0,AC11/AB11,"")', '=IF(AB12>0,AC12/AB12,"")', '=IF(AB13>0,AC13/AB13,"")', '=IF(AB14>0,AC14/AB14,"")', '=IF(AB15>0,AC15/AB15,"")', '=IF(AB16>0,AC16/AB16,"")', '=IF(AB17>0,AC17/AB17,"")', '=IF(AB18>0,AC18/AB18,"")', '=IF(AB19>0,AC19/AB19,"")', '=IF(AB20>0,AC20/AB20,"")', '=IF(AB21>0,AC21/AB21,"")', '=IF(AB22>0,AC22/AB22,"")', '=IF(AB23>0,AC23/AB23,"")']

sheet_names = wb.sheetnames
x=4
j=1
#sheet1.cell(row=4,column=14).value = "=IF(L"+str(i[0])+">0,M"+"4"+"/L"+"4"+","")"
while j<32:
	sheet1 = wb[sheet_names[j]]
	#print(sheet1)
	for i in range(0,20):
		sheet1['G'+str(x)] = data7[i]
		sheet1['H'+str(x)] = data8[i]
		sheet1['I'+str(x)] = data9[i]
		sheet1['J'+str(x)] = data10[i]
		sheet1['N'+str(x)] = data14[i]
		sheet1['R'+str(x)] = data18[i]
		sheet1['V'+str(x)] = data22[i]
		sheet1['Z'+str(x)] = data26[i]
		sheet1['AD'+str(x)] = data30[i]
		x += 1
	j += 1
	x=4

sheet2 = wb1.get_sheet_by_name('个人汇总')
all_col = []
for row in sheet2.iter_rows(max_row=5, min_row=5, max_col=163, min_col=4):
	for cell in row:
		all_col.append(cell.value)
sheet3 = wb.get_sheet_by_name('个人汇总')
index = 0
for k in range(4,164):
	sheet3.cell(row=5,column=k).value = all_col[index]
	index += 1

#print(sheet_names)

wb.save(r'E:\lianxi\PY\12.xlsx')



